from glob import glob
from openpyxl import Workbook
from pathlib import WindowsPath
import xlrd
import argparse
import re

def combine_sheets(files,output_file):
    out = Workbook()
    out.remove(out.active)
    panel_expr = re.compile(r'Panel\s*\d+\s*[A-Za-z]i*')
    for file in files:
        wb = xlrd.open_workbook(file)
        sheets = wb.sheet_names()
        ws = wb.sheet_by_index(0) 
        name = panel_expr.findall(file) 
        if name:
            out_sheet = out.create_sheet(name[0])
            name = name[0]
        else:
            name = WindowsPath(file).stem
            out_sheet = out.create_sheet(name)
        for row in range(ws.nrows):
            out_sheet.append([i.value for i in ws.row(row)])
    out.save(output_file)
            


def main():
    """Command Line Script"""
    # Prepare to accept command line arguments
    parser = argparse.ArgumentParser(
        description="""Combine multiple excel files with a single excel"""\
            """spreadsheet into one excel file with multiple sheets."""
    )
    parser.add_argument(
        '-i', 
        '--input_path',
        help='File path where excel files are located, can be folder or'\
            ' individual file'
    )
    parser.add_argument(
        '-o',
        '--output_file',
        help='File path to output excel file'
    )
    parser.add_argument(
        '-g',
        '--glob',
        action='store_true',
        help='Use glob to find files specified by input_path'
    )
    parser.add_argument(
        '-r',
        '--recursive',
        action='store_true',
        help='Perform glob search recursively'
    )

    args = parser.parse_args()
    # handle flags relating to glob usage for file searching
    # This allows the user to search for files using glob directly
    if args.glob:
        files = glob(args.input_path,recursive=args.recursive)
        data = combine_sheets(files,args.output_file)
    else:
        path = WindowsPath(args.input_path)
        data = combine_sheets([str(p) for p in path.glob('*xls*')],args.output_file)
    print(
        f'Successfully created output {args.output_file}'
    )
